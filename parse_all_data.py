import openpyxl
import os
import json

base_dir = "/Users/jose/.gemini/antigravity/scratch/porra-mundial"
files = [
    "Luismi.xlsx", "Nacho.xlsx", "Mikel.xlsx", "PiliP.xlsx", "Lidia.xlsx", 
    "Javi.xlsx", "Camina.xlsx", "Andrés.xlsx", "Alber.xlsx", "Iván.xlsx", "8pi.xlsx"
]

def clean_name(filename):
    return filename.replace(".xlsx", "")

# 1. Load a sample workbook to extract official fixtures and groups
ref_file = os.path.join(base_dir, "PiliP.xlsx")
wb_ref = openpyxl.load_workbook(ref_file, data_only=True)

# Extract official teams and groups
# In WORLDCUP sheet, col BD is team, col BC (column 55) is group
sheet_wc = wb_ref["WORLDCUP"]
groups_data = {}
groups = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
for g_idx, g_name in enumerate(groups):
    start_row = 5 + 8 * g_idx
    groups_data[g_name] = []
    for r in range(start_row, start_row + 4):
        team = sheet_wc.cell(row=r, column=1).value # Col A (1) has the official team names
        if team:
            groups_data[g_name].append(team)

# Extract official group fixtures (72 matches)
fixtures = []
for g_idx, g_name in enumerate(groups):
    start_row = 4 + 8 * g_idx
    for m in range(6):
        r = start_row + m
        home = sheet_wc.cell(row=r, column=27).value # Col AA
        away = sheet_wc.cell(row=r, column=32).value # Col AF
        date_val = sheet_wc.cell(row=r, column=24).value # Col X
        jor = sheet_wc.cell(row=r, column=26).value # Col Z
        match_num = sheet_wc.cell(row=r, column=34).value # Col AH
        
        date_str = ""
        if date_val:
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d %H:%M")
            else:
                date_str = str(date_val)
        
        fixtures.append({
            "matchNum": match_num,
            "group": g_name,
            "home": home,
            "away": away,
            "date": date_str,
            "jor": jor
        })

# 2. Extract Combinaciones
sheet_comb = wb_ref["Combinaciones"]
combinaciones = {}
for r in range(2, 500):
    code = sheet_comb.cell(row=r, column=14).value # Column N is 14
    if not code:
        continue
    letters = []
    for c in range(16, 24): # P to W
        val = sheet_comb.cell(row=r, column=c).value
        if val:
            letters.append(val[-1])
        else:
            letters.append("?")
    combinaciones[code] = "".join(letters)

# 3. Extract participant predictions
predictions = {}
contenders = []

for f in files:
    name = clean_name(f)
    contenders.append(name)
    print(f"Parsing predictions for {name}...")
    filepath = os.path.join(base_dir, f)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Read Pool sheet
    sheet_pool = wb["Pool"]
    
    # Parse group matches predictions (rows 6 to 77)
    group_matches_preds = {}
    for r in range(6, 78):
        matchup = sheet_pool.cell(row=r, column=2).value
        pred = sheet_pool.cell(row=r, column=3).value
        
        winner = "1"
        gh = 0
        ga = 0
        if pred and "|" in pred:
            parts = pred.split("|")
            winner = parts[0]
            score_part = parts[1]
            if "-" in score_part:
                scores = score_part.split("-")
                try:
                    gh = int(scores[0])
                    ga = int(scores[1])
                except ValueError:
                    pass
        
        home, away = matchup.split("-")
        
        # Find matching fixture to get matchNum
        matched_num = None
        for fix in fixtures:
            if fix["home"] == home and fix["away"] == away:
                matched_num = fix["matchNum"]
                break
        
        if matched_num is not None:
            group_matches_preds[str(matched_num)] = {
                "winner": winner,
                "gh": gh,
                "ga": ga
            }
        else:
            print(f"ERROR: Could not find fixture for {home} vs {away}")

    
    # Read predicted group standings from WORLDCUP sheet
    sheet_wc_p = wb["WORLDCUP"]
    group_rankings_preds = {}
    for g_idx, g_name in enumerate(groups):
        start_row = 6 + 8 * g_idx
        group_rankings_preds[g_name] = []
        for r in range(start_row, start_row + 4):
            team = sheet_wc_p.cell(row=r, column=38).value # Col AL (38) is Selección
            if team:
                group_rankings_preds[g_name].append(team)
    
    # Parse qualifiers
    # Octavos (1/8): rows 182-197
    octavos = [sheet_pool.cell(row=r, column=3).value for r in range(182, 198)]
    # Cuartos (1/4): rows 210-217
    cuartos = [sheet_pool.cell(row=r, column=3).value for r in range(210, 218)]
    # Semis (1/2): rows 226-229
    semis = [sheet_pool.cell(row=r, column=3).value for r in range(226, 230)]
    # Finalistas: rows 240-241
    finalistas = [sheet_pool.cell(row=r, column=3).value for r in range(240, 242)]
    # 3-4 Match participants: rows 236-237
    match34 = [sheet_pool.cell(row=r, column=3).value for r in range(236, 238)]
    
    # Cuadro de honor
    champion = sheet_pool.cell(row=250, column=3).value
    subchampion = sheet_pool.cell(row=251, column=3).value
    third_place = sheet_pool.cell(row=252, column=3).value
    pichichi = sheet_pool.cell(row=253, column=3).value
    mvp = sheet_pool.cell(row=256, column=3).value
    
    # Parse knockout match score predictions
    ko_match_preds = {}
    # Row index to match key mappings
    # 1/16 matches: 164-179 -> Matches 73-88
    for idx, r in enumerate(range(164, 180)):
        match_num = 73 + idx
        val_c = sheet_pool.cell(row=r, column=3).value
        ko_match_preds[str(match_num)] = val_c
        
    # 1/8 matches: 200-207 -> Matches 89-96
    for idx, r in enumerate(range(200, 208)):
        match_num = 89 + idx
        val_c = sheet_pool.cell(row=r, column=3).value
        ko_match_preds[str(match_num)] = val_c
        
    # 1/4 matches: 220-223 -> Matches 97-100
    for idx, r in enumerate(range(220, 224)):
        match_num = 97 + idx
        val_c = sheet_pool.cell(row=r, column=3).value
        ko_match_preds[str(match_num)] = val_c
        
    # Semis: 232-233 -> Matches 101-102
    for idx, r in enumerate(range(232, 234)):
        match_num = 101 + idx
        val_c = sheet_pool.cell(row=r, column=3).value
        ko_match_preds[str(match_num)] = val_c
        
    # 3-4 match: 244 -> Match 103
    ko_match_preds["103"] = sheet_pool.cell(row=244, column=3).value
    
    # Final match: 247 -> Match 104
    ko_match_preds["104"] = sheet_pool.cell(row=247, column=3).value

    # Parse penalty shootout predictions from WORLDCUP sheet
    ko_penalties = {}
    for r in range(1, 200):
        match_num_val = sheet_wc_p.cell(row=r, column=34).value # Col AH (34)
        if isinstance(match_num_val, int) and 73 <= match_num_val <= 104:
            home_pen = sheet_wc_p.cell(row=r, column=28).value # Col AB (28)
            away_pen = sheet_wc_p.cell(row=r, column=31).value # Col AE (31)
            if home_pen is not None or away_pen is not None:
                try:
                    hp = int(home_pen) if home_pen is not None and str(home_pen).strip() != "" else 0
                    ap = int(away_pen) if away_pen is not None and str(away_pen).strip() != "" else 0
                    ko_penalties[str(match_num_val)] = {"home": hp, "away": ap}
                except ValueError:
                    pass

    predictions[name] = {
        "groupMatches": group_matches_preds,
        "groupRankings": group_rankings_preds,
        "octavos": octavos,
        "cuartos": cuartos,
        "semis": semis,
        "finalistas": finalistas,
        "match34": match34,
        "thirdPlace": third_place,
        "champion": champion,
        "subchampion": subchampion,
        "pichichi": pichichi,
        "mvp": mvp,
        "koMatches": ko_match_preds,
        "koPenalties": ko_penalties
    }

# 4. Generate the JS file contents
js_content = f"""// Generated World Cup Sweepstakes Data
const CONTENDERS = {json.dumps(contenders, indent=2, ensure_ascii=False)};

const GROUPS_DATA = {json.dumps(groups_data, indent=2, ensure_ascii=False)};

const FIXTURES = {json.dumps(fixtures, indent=2, ensure_ascii=False)};

const COMBINATIONS = {json.dumps(combinaciones, indent=2, ensure_ascii=False)};

const PREDICTIONS = {json.dumps(predictions, indent=2, ensure_ascii=False)};
"""

# Write to data.js
js_filepath = os.path.join(base_dir, "data.js")
with open(js_filepath, "w", encoding="utf-8") as f_out:
    f_out.write(js_content)

print(f"Data parsed successfully and written to {js_filepath}!")
print(f"Contenders: {contenders}")
